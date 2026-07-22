import os
import webview
import openpyxl

class ReporteIniController:
    def __init__(self):
        pass

    def generar_excel_desde_plantilla(self, datos_estudiantes, nombre_archivo):
        try:
            # Apuntamos a la raíz del backend (subimos 3 niveles: excels -> controllers -> BackEnd)
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            
            # Buscamos específicamente el archivo matricula_inicial.xlsx en la carpeta Excels
            ruta_plantilla = os.path.join(base_dir, "Excels", "matricula_inicial.xlsx")
            
            if not os.path.exists(ruta_plantilla):
                return {
                    "status": "error", 
                    "message": f"No se encontró el archivo en la ruta esperada: {ruta_plantilla}"
                }

            window = webview.windows[0]
            result = window.create_file_dialog(
                webview.SAVE_DIALOG,
                directory='',
                save_filename=f"{nombre_archivo}.xlsx",
                file_types=('Archivos Excel (*.xlsx)', 'Todos los archivos (*.*)')
            )

            if not result:
                return {"status": "cancelado"}

            ruta_guardado = result[0]

            # Cargar plantilla preservando estilos e imágenes
            wb = openpyxl.load_workbook(ruta_plantilla)
            ws = wb.active

            # =========================================================
            # FUNCIÓN PARA ESCRITURA EN CELDAS COMBINADAS
            # =========================================================
            def escribir_celda_segura(fila, col, valor):
                coordenada = ws.cell(row=fila, column=col).coordinate
                for rango in ws.merged_cells.ranges:
                    if coordenada in rango:
                        # Escribe SIEMPRE en la esquina superior izquierda del rango combinado
                        ws.cell(row=rango.min_row, column=rango.min_col).value = valor
                        return
                # Si no está combinada, escribe de forma estándar
                ws.cell(row=fila, column=col).value = valor

            # =========================================================
            # ENCABEZADOS DE LA BARRA SUPERIOR (Fila 10)
            # =========================================================
            cantidad_alumnos = len(datos_estudiantes)
            grado = ""
            seccion = ""
            
            if cantidad_alumnos > 0:
                grado = str(datos_estudiantes[0].get('nivelEstudio') or datos_estudiantes[0].get('grado') or '').upper()
                seccion = str(datos_estudiantes[0].get('seccion', '')).upper()

            # Fila 10 A-C (Columna 1) - Siempre incluir la palabra "Grado: "
            texto_grado = f"Grado: {grado}".strip()
            escribir_celda_segura(10, 1, texto_grado)
            
            # Fila 10 F (Columna 6) - Sección
            escribir_celda_segura(10, 6, seccion)

            # Fila 10 L (Columna 12) - Cantidad de alumnos en el archivo
            escribir_celda_segura(10, 12, cantidad_alumnos)

            # Fila 10 S (Columna 19) - Cantidad de alumnos en esta página
            escribir_celda_segura(10, 19, cantidad_alumnos)

            # =========================================================
            # COORDENADAS EXACTAS DE ESTUDIANTES (Fila 14 en adelante)
            # =========================================================
            fila_actual = 14 
            limite_filas = 48 

            for i, est in enumerate(datos_estudiantes):
                if fila_actual > limite_filas:
                    print("⚠️ Se alcanzó el límite de estudiantes por plantilla (35).")
                    break
                
                # Número de lista (Opcional, en la A)
                escribir_celda_segura(fila_actual, 1, str(i + 1).zfill(2))
                
                # Fila 14 B-C (Columna 2) - Cédula Escolar
                cedula = est.get('cedulaEscolar') or est.get('cedula_estudiantil') or ''
                escribir_celda_segura(fila_actual, 2, cedula)

                # Fila 14 D-H (Columna 4) - Apellidos
                apellido = est.get('est_apellido') or est.get('apellido') or est.get('apellidos') or ''
                escribir_celda_segura(fila_actual, 4, str(apellido).upper())
                
                # Fila 14 I-M (Columna 9) - Nombres
                nombre = est.get('est_nombre') or est.get('nombre') or est.get('nombres') or ''
                escribir_celda_segura(fila_actual, 9, str(nombre).upper())

                # Fila 14 N (Columna 14) - Sexo
                genero = est.get('est_genero') or est.get('genero') or ''
                escribir_celda_segura(fila_actual, 14, genero[0].upper() if genero else '')

                # Fila 14 O, P, Q - Fecha de Nacimiento
                fecha_nac = est.get('est_fecha_nacimiento') or est.get('fechaNacimiento') or ''
                if fecha_nac:
                    sep = '-' if '-' in fecha_nac else '/'
                    partes = fecha_nac.split(sep)
                    if len(partes) == 3:
                        if len(partes[0]) == 4: 
                            dia, mes, anio = partes[2], partes[1], partes[0]
                        else:
                            dia, mes, anio = partes[0], partes[1], partes[2]
                            
                        # Fila 14 O (Columna 15) - Día
                        escribir_celda_segura(fila_actual, 15, dia)  
                        # Fila 14 P (Columna 16) - Mes
                        escribir_celda_segura(fila_actual, 16, mes)  
                        # Fila 14 Q (Columna 17) - Año
                        escribir_celda_segura(fila_actual, 17, anio) 

                # Avanzamos al siguiente registro
                fila_actual += 1

            wb.save(ruta_guardado)
            return {"status": "success"}

        except Exception as e:
            print(f"❌ Error al generar Excel inicial: {e}")
            return {"status": "error", "message": str(e)}