import os
import webview
import openpyxl

class ReporteController:
    def __init__(self):
        pass

    def generar_excel_desde_plantilla(self, datos_estudiantes, nombre_archivo):
        try:
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            ruta_plantilla = os.path.join(base_dir, "Excels", "plantilla_matricula.xlsx")
            
            if not os.path.exists(ruta_plantilla):
                return {
                    "status": "error", 
                    "message": f"No se encontró el archivo plantilla_matricula.xlsx en la ruta esperada: {ruta_plantilla}"
                }

            window = webview.windows[0]
            result = window.create_file_dialog(
                webview.SAVE_DIALOG,
                directory='',
                save_filename=f"{nombre_archivo}.xlsx",
                file_types=('Archivos Excel (*.xlsx)', 'Todos los archivos (*.*)')
            )

            if not result:
                return {"status": "cancelado"}

            ruta_guardado = result[0]

            wb = openpyxl.load_workbook(ruta_plantilla)
            ws = wb.active

            # =========================================================
            # FUNCIÓN PARA ESCRITURA EN CELDAS COMBINADAS
            # =========================================================
            def escribir_celda_segura(fila, col, valor):
                coordenada = ws.cell(row=fila, column=col).coordinate
                for rango in ws.merged_cells.ranges:
                    if coordenada in rango:
                        # Escribe SIEMPRE en la esquina superior izquierda del rango combinado
                        ws.cell(row=rango.min_row, column=rango.min_col).value = valor
                        return
                # Si no está combinada, escribe de forma estándar
                ws.cell(row=fila, column=col).value = valor

            # =========================================================
            # ENCABEZADOS DE LA BARRA SUPERIOR (Fila 13)
            # =========================================================
            cantidad_alumnos = len(datos_estudiantes)

            # L (12) / M (13) - N° de Estudiantes en la Sección
            escribir_celda_segura(13, 12, cantidad_alumnos)             
            
            # T (20) / U (21) - N° de Estudiantes en esta página
            escribir_celda_segura(13, 20, cantidad_alumnos) 

            # =========================================================
            # ITERACIÓN Y LLENADO DE ESTUDIANTES (DOS TABLAS)
            # =========================================================
            fila_superior = 18 
            fila_inferior = 40 

            for i, est in enumerate(datos_estudiantes):
                # ---------------------------------------------------------
                # TABLA SUPERIOR (Fila 18 en adelante)
                # ---------------------------------------------------------
                
                # N° de lista
                escribir_celda_segura(fila_superior, 1, i + 1)
                
                # Cédula (Columna B / 2)
                cedula = est.get('cedulaEscolar') or est.get('cedula_estudiantil') or ''
                escribir_celda_segura(fila_superior, 2, cedula)
                
                # Lugar de Residencia / Nacimiento (De la E a la I -> Origen E / 5)
                direccion = est.get('est_direccion') or est.get('direccion') or ''
                escribir_celda_segura(fila_superior, 5, str(direccion).upper())
                
                # Sexo (Columna K / 11)
                genero = est.get('est_genero') or est.get('genero') or ''
                escribir_celda_segura(fila_superior, 11, genero[0].upper() if genero else '')

                # Fecha de Nacimiento: Día L/12, Mes M/13, Año N/14
                fecha_nac = est.get('est_fecha_nacimiento') or est.get('fechaNacimiento') or ''
                if fecha_nac:
                    sep = '-' if '-' in fecha_nac else '/'
                    partes = fecha_nac.split(sep)
                    if len(partes) == 3:
                        if len(partes[0]) == 4: 
                            dia, mes, anio = partes[2], partes[1], partes[0]
                        else:
                            dia, mes, anio = partes[0], partes[1], partes[2]
                            
                        escribir_celda_segura(fila_superior, 12, dia)  
                        escribir_celda_segura(fila_superior, 13, mes)  
                        escribir_celda_segura(fila_superior, 14, anio) 

                # ---------------------------------------------------------
                # TABLA INFERIOR (Fila 40 en adelante)
                # ---------------------------------------------------------
                
                # N° de lista inferior
                escribir_celda_segura(fila_inferior, 1, f"{i + 1:02d}")
                
                # Apellidos (De la columna B / 2 hasta la J)
                apellido = est.get('est_apellido') or est.get('apellido') or est.get('apellidos') or ''
                escribir_celda_segura(fila_inferior, 2, str(apellido).upper())

                # Nombres (De la columna K / 11 hasta la U)
                nombre = est.get('est_nombre') or est.get('nombre') or est.get('nombres') or ''
                escribir_celda_segura(fila_inferior, 11, str(nombre).upper())

                # Avanzar filas
                fila_superior += 1
                fila_inferior += 1

            wb.save(ruta_guardado)
            return {"status": "success"}

        except Exception as e:
            print(f"❌ Error al generar Excel: {e}")
            return {"status": "error", "message": str(e)}