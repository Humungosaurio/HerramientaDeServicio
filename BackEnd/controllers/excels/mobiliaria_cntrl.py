import os
import webview
import openpyxl

class MobiliariaController:
    def __init__(self):
        pass

    def generar_excel_desde_plantilla(self, datos_mobiliario, nombre_archivo):
        try:
            # Subimos 3 niveles si está en backend/controllers/excels/ para llegar a la raíz
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            ruta_plantilla = os.path.join(base_dir, "Excels", "mobiliaria.xlsx")
            
            # Fallback
            if not os.path.exists(ruta_plantilla):
                ruta_plantilla = os.path.join(os.path.dirname(os.path.dirname(__file__)), "mobiliaria.xlsx")

            if not os.path.exists(ruta_plantilla):
                return {"status": "error", "message": f"No se encontró la plantilla mobiliaria.xlsx en: {ruta_plantilla}"}

            window = webview.windows[0]
            result = window.create_file_dialog(
                webview.SAVE_DIALOG,
                directory='',
                save_filename=f"{nombre_archivo}.xlsx",
                file_types=('Archivos Excel (*.xlsx)', 'Todos los archivos (*.*)')
            )

            if not result:
                return {"status": "cancelado"}

            ruta_guardado = result[0]

            # Cargar plantilla
            wb = openpyxl.load_workbook(ruta_plantilla)
            ws = wb.active

            # =========================================================
            # FILA DONDE EMPIEZAN LOS DATOS (Fila 4 según tu imagen)
            # =========================================================
            fila_actual = 4 

            for item in datos_mobiliario:
                # Columna A (1): Nombre
                ws.cell(row=fila_actual, column=1, value=item.get('nombre', '')) 
                
                # Columna B (2): Descripción (Usamos las observaciones del frontend)
                ws.cell(row=fila_actual, column=2, value=item.get('observaciones', ''))
                
                # Columna C (3): uso
                estado = "En uso" if item.get('enUso') else "En desuso"
                ws.cell(row=fila_actual, column=3, value=estado)
                
                # Columna D (4): Cantidad
                ws.cell(row=fila_actual, column=4, value=item.get('cantidad', 0))
                
                fila_actual += 1

            wb.save(ruta_guardado)
            return {"status": "success"}

        except Exception as e:
            print(f"❌ Error al generar Excel de mobiliaria: {e}")
            return {"status": "error", "message": str(e)}