import os
import sqlite3
import webview
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

class Asis_Totales_Controller:
    def __init__(self, db_path):
        self.db_path = db_path

    def calcular_edad(self, fecha_nacimiento):
        if not fecha_nacimiento:
            return 0
            
        fecha_str = str(fecha_nacimiento).strip()
        fn = None
        
        formatos = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%y', '%d-%m-%y']
        
        for fmt in formatos:
            try:
                fn = datetime.strptime(fecha_str, fmt)
                if fn.year < 100: 
                    fn = fn.replace(year=fn.year + 2000)
                break
            except ValueError:
                continue
                
        if not fn:
            return 0
            
        hoy = datetime.now()
        edad = hoy.year - fn.year - ((hoy.month, hoy.day) < (fn.month, fn.day))
        
        if edad < 0 or edad > 100:
            return 0
            
        return edad

    def generar_excel_asistencias_totales(self, parametros):
        try:
            mes = str(parametros.get('mes') or 'Julio').strip()
            grado = str(parametros.get('grado') or '').strip()
            turno = str(parametros.get('turno') or '').strip()
            seccion = str(parametros.get('seccion') or '').strip()
            
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            ruta_plantilla = os.path.join(base_dir, "Excels", "Asistencias_totales.xlsx")
            
            if not os.path.exists(ruta_plantilla):
                return {"status": "error", "message": f"No se encontró la plantilla en: {ruta_plantilla}"}

            wb = openpyxl.load_workbook(ruta_plantilla)
            ws = wb.active
            ws.title = f"Resumen {mes}"

            center_align = Alignment(horizontal="center", vertical="center")

            def escribir_seguro(celda, valor):
                try:
                    ws[celda] = valor
                    ws[celda].alignment = center_align
                except AttributeError:
                    pass

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("""
                SELECT salon_id FROM salones 
                WHERE grado = ? AND turno = ? AND seccion = ?
            """, (grado, turno, seccion))
            salon_row = cursor.fetchone()
            
            if not salon_row:
                conn.close()
                return {"status": "error", "message": "No se encontró el salón especificado."}
            
            salon_id = salon_row[0]

            cursor.execute("""
                SELECT cedula_estudiantil, est_fecha_nacimiento, est_genero 
                FROM Estudiante 
                WHERE salon_id = ? AND (LOWER(estado) != 'retirado' OR estado IS NULL)
            """, (salon_id,))
            estudiantes = cursor.fetchall()

            if not estudiantes:
                conn.close()
                return {"status": "error", "message": "No hay estudiantes vigentes en esta sección."}

            ids_alumnos = [str(est[0]) for est in estudiantes]
            
            # --- CÁLCULO DE EDADES Y SEXO ---
            conteo_edades = {}
            for est in estudiantes:
                edad = self.calcular_edad(est[1])
                genero = str(est[2]).strip().lower()
                
                if edad not in conteo_edades:
                    conteo_edades[edad] = {'V': 0, 'H': 0}
                    
                if genero in ['masculino', 'm', 'v', 'varon', 'varón', 'niño']:
                    conteo_edades[edad]['V'] += 1
                elif genero in ['femenino', 'f', 'h', 'hembra', 'niña']:
                    conteo_edades[edad]['H'] += 1

            # --- ESCRIBIR EDADES --- 
            fila_edad = 6
            total_v_edades = 0
            total_h_edades = 0
            
            for edad in sorted(conteo_edades.keys()):
                if fila_edad > 9:
                    break
                v_count = conteo_edades[edad]['V']
                h_count = conteo_edades[edad]['H']
                t_count = v_count + h_count
                
                escribir_seguro(f'G{fila_edad}', f"{edad} Años" if edad > 0 else "S/E")
                escribir_seguro(f'H{fila_edad}', v_count)
                escribir_seguro(f'I{fila_edad}', h_count)
                escribir_seguro(f'J{fila_edad}', t_count)
                
                total_v_edades += v_count
                total_h_edades += h_count
                fila_edad += 1

            # --- ESCRIBIR TOTALES DE EDADES ---
            escribir_seguro('H10', total_v_edades)
            escribir_seguro('I10', total_h_edades)
            escribir_seguro('J10', total_v_edades + total_h_edades)

            # --- CONSULTA DE ASISTENCIAS ---
            placeholders = ','.join('?' * len(ids_alumnos))
            cursor.execute(f"""
                SELECT a.semana, a.dia_semana, e.est_genero, COUNT(*)
                FROM asistencias a
                JOIN Estudiante e ON a.cedula_estudiantil = e.cedula_estudiantil
                WHERE a.mes = ? AND a.estado = 'Presente' AND a.cedula_estudiantil IN ({placeholders})
                GROUP BY a.semana, a.dia_semana, e.est_genero
            """, [mes] + ids_alumnos)
            
            asistencias_db = cursor.fetchall()
            conn.close()

            asistencia_procesada = {}
            for row in asistencias_db:
                semana_db = row[0].strip()
                dia_db = row[1].lower().strip()
                genero_db = str(row[2]).strip().lower()
                cantidad = row[3]

                if semana_db not in asistencia_procesada:
                    asistencia_procesada[semana_db] = {}
                if dia_db not in asistencia_procesada[semana_db]:
                    asistencia_procesada[semana_db][dia_db] = {'V': 0, 'H': 0}
                
                if genero_db in ['masculino', 'm', 'v', 'varon', 'varón', 'niño']:
                    asistencia_procesada[semana_db][dia_db]['V'] += cantidad
                elif genero_db in ['femenino', 'f', 'h', 'hembra', 'niña']:
                    asistencia_procesada[semana_db][dia_db]['H'] += cantidad

            # --- ESCRIBIR METADATOS --- 
            escribir_seguro('I14', f"{grado} {turno} - Sec. {seccion}")
            escribir_seguro('I15', mes.capitalize())

            # --- MAPEO DE FILAS ---
            mapa_filas_dias = {
                "Semana 1": {"lunes": 3, "martes": 4, "miércoles": 5, "miercoles": 5, "jueves": 6, "viernes": 7},
                "Semana 2": {"lunes": 8, "martes": 9, "miércoles": 10, "miercoles": 10, "jueves": 11, "viernes": 12},
                "Semana 3": {"lunes": 13, "martes": 14, "miércoles": 15, "miercoles": 15, "jueves": 16, "viernes": 17},
                "Semana 4": {"lunes": 18, "martes": 19, "miércoles": 20, "miercoles": 20, "jueves": 21, "viernes": 22}
            }

            for semana, dias in asistencia_procesada.items():
                if semana in mapa_filas_dias:
                    for dia, conteos in dias.items():
                        dia_sin_acento = dia.replace('é', 'e')
                        if dia in mapa_filas_dias[semana] or dia_sin_acento in mapa_filas_dias[semana]:
                            fila = mapa_filas_dias[semana].get(dia) or mapa_filas_dias[semana].get(dia_sin_acento)
                            
                            v_asist = conteos['V']
                            h_asist = conteos['H']
                            t_asist = v_asist + h_asist
                            
                            escribir_seguro(f'C{fila}', v_asist)
                            escribir_seguro(f'D{fila}', h_asist)
                            escribir_seguro(f'E{fila}', t_asist)

            # --- CÁLCULO Y LLENADO DE FILAS INFERIORES (23, 24 y 25) ---
            sum_v = sum(conteos['V'] for sem in asistencia_procesada.values() for conteos in sem.values())
            sum_h = sum(conteos['H'] for sem in asistencia_procesada.values() for conteos in sem.values())
            sum_total = sum_v + sum_h

            # Fila 23: Total de asistencias
            escribir_seguro('C23', sum_v)
            escribir_seguro('D23', sum_h)
            escribir_seguro('E23', sum_total)

            # Fila 24: Promedio de asistencias (sobre 20 días hábiles del mes)
            prom_v = round(sum_v / 20.0, 2) if sum_v > 0 else 0.0
            prom_h = round(sum_h / 20.0, 2) if sum_h > 0 else 0.0
            prom_total = round(sum_total / 20.0, 2) if sum_total > 0 else 0.0

            escribir_seguro('C24', prom_v)
            escribir_seguro('D24', prom_h)
            escribir_seguro('E24', prom_total)

            # Fila 25: Porcentaje de asistencias
            max_v = total_v_edades * 20
            max_h = total_h_edades * 20
            max_total = len(estudiantes) * 20

            pct_v = round((sum_v / max_v) * 100, 1) if max_v > 0 else 0.0
            pct_h = round((sum_h / max_h) * 100, 1) if max_h > 0 else 0.0
            pct_total = round((sum_total / max_total) * 100, 1) if max_total > 0 else 0.0

            escribir_seguro('C25', f"{pct_v}%" if max_v > 0 else "0%")
            escribir_seguro('D25', f"{pct_h}%" if max_h > 0 else "0%")
            escribir_seguro('E25', f"{pct_total}%" if max_total > 0 else "0%")

            # --- GUARDADO ---
            try:
                window = webview.windows[0]
                
                # MODIFICADO: Generamos el nombre limpiando espacios y evitando duplicar la sección
                grado_limpio = grado.replace(" ", "_")
                seccion_limpia = seccion.replace(" ", "_")
                nombre_sugerido = f"Asistencias_Totales_{grado_limpio}_Sec_{seccion_limpia}_{mes}.xlsx"
                nombre_sugerido = nombre_sugerido.replace("/", "-").replace("\\", "-")
                
                result = window.create_file_dialog(
                    webview.SAVE_DIALOG, 
                    directory='', 
                    save_filename=nombre_sugerido
                )

                if result:
                    ruta_guardado = result[0] if isinstance(result, tuple) else result
                    if not ruta_guardado.endswith('.xlsx'):
                        ruta_guardado += '.xlsx'
                    wb.save(ruta_guardado)
                    return {"status": "success", "ruta": ruta_guardado}
                else:
                    return {"status": "error", "message": "Guardado cancelado por el usuario."}
                    
            except Exception as w_err:
                return {"status": "error", "message": f"Error abriendo ventana de guardado: {str(w_err)}"}

        except Exception as e:
            return {"status": "error", "message": str(e)}