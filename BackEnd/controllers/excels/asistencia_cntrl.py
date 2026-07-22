import os
import sqlite3
import webview
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

class AsistenciasController:
    def __init__(self, db_path):
        self.db_path = db_path

    def generar_excel_asistencias(self, parametros):
        try:
            datos_alumnos = parametros.get('datos', [])
            mes = str(parametros.get('mes') or 'Julio').strip()
            
            # 1. Leer el filtro enviado desde React
            opcion_filtro = str(parametros.get('opcion_filtro') or '').strip()

            if not datos_alumnos:
                return {"status": "error", "message": "No hay datos para exportar."}

            # 2. Parsear Grado (Nivel), Sección y Turno a partir del texto del menú
            grado = ""
            seccion = ""
            turno = ""

            if " - Sec. " in opcion_filtro:
                partes = opcion_filtro.split(" - Sec. ")
                seccion = partes[1].strip()
                if "Mañana" in partes[0]:
                    turno = "MAÑANA"
                    grado = partes[0].replace("Mañana", "").strip()
                elif "Tarde" in partes[0]:
                    turno = "TARDE"
                    grado = partes[0].replace("Tarde", "").strip()
                else:
                    grado = partes[0].strip()
            else:
                grado = str(parametros.get('grado') or opcion_filtro).strip()
                seccion = str(parametros.get('seccion') or '').strip()
                turno = str(parametros.get('turno') or '').strip()

            if not seccion:
                seccion = "General"
            if not grado:
                grado = "General"

            # 3. Cargar la plantilla de Excel preexistente
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
            ruta_plantilla = os.path.join(base_dir, "Excels", "Asistencias_detalladas.xlsx")
            
            if not os.path.exists(ruta_plantilla):
                return {"status": "error", "message": f"No se encontró la plantilla en: {ruta_plantilla}"}

            wb = openpyxl.load_workbook(ruta_plantilla)
            ws = wb.active
            ws.title = f"Asistencias {mes}"

            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            
            # Estilos para nueva columna de Porcentaje
            negrita = Font(bold=True)
            borde_fino = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            # 4. Llenar datos de la cabecera (Fila 2) con SECCIÓN específica y explícita
            if type(ws['A2']).__name__ != 'MergedCell':
                ws['A2'] = f"MES: {mes.upper()}    NIVEL: {grado.upper()}    SECCIÓN: {seccion.upper()}    TURNO: {turno.upper()}"

            # 5. Consultar toda la asistencia del mes desde la base de datos SQLite
            ids_alumnos = [str(al['id']) for al in datos_alumnos]
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            placeholders = ','.join('?' * len(ids_alumnos))
            query = f"""
                SELECT cedula_estudiantil, semana, dia_semana, estado
                FROM asistencias
                WHERE mes = ? AND cedula_estudiantil IN ({placeholders})
            """
            cursor.execute(query, [mes] + ids_alumnos)
            asistencias_db = cursor.fetchall()
            conn.close()

            # Mapear matriz temporal de asistencias en memoria
            mapa_asis = {}
            for row in asistencias_db:
                ced = str(row[0])
                semana = str(row[1]).strip()
                dia = str(row[2]).strip().capitalize()
                estado = row[3]
                
                if ced not in mapa_asis:
                    mapa_asis[ced] = {}
                if semana not in mapa_asis[ced]:
                    mapa_asis[ced][semana] = {}
                    
                mapa_asis[ced][semana][dia] = (estado == 'Presente')

            # 6. Rellenar datos en la plantilla (A partir de la fila 4)
            semanas_cols = {
                "Semana 1": 4,  # Inicia en Columna D
                "Semana 2": 9,  # Inicia en Columna I
                "Semana 3": 14, # Inicia en Columna N
                "Semana 4": 19  # Inicia en Columna S
            }
            dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
            total_dias_mes = len(semanas_cols) * len(dias)  # 20 días hábiles en las 4 semanas

            # Configurar cabecera para Porcentaje de Asistencia (Columna 25 / Y)
            celda_cabecera_porc = ws.cell(row=3, column=25)
            celda_cabecera_porc.value = "% ASIST."
            celda_cabecera_porc.font = negrita
            celda_cabecera_porc.alignment = center_align
            celda_cabecera_porc.border = borde_fino
            ws.column_dimensions['Y'].width = 12

            # Contadores de género
            total_varones = 0
            total_hembras = 0

            fila_actual = 4
            for i, al in enumerate(datos_alumnos, start=1):
                ced = str(al['id'])
                nombre = al.get('nombre', '')
                
                # Conteo de Varones y Hembras y asignación de la letra (V o H)
                genero = str(al.get('genero') or al.get('sexo') or '').strip().upper()
                letra_sexo = ""
                if genero in ['M', 'MASCULINO', 'V', 'VARON', 'VARÓN']:
                    total_varones += 1
                    letra_sexo = "V"
                elif genero in ['F', 'FEMENINO', 'H', 'HEMBRA']:
                    total_hembras += 1
                    letra_sexo = "H"

                # Número de lista (Columna 1 / A)
                celda_num = ws.cell(row=fila_actual, column=1)
                if type(celda_num).__name__ != 'MergedCell':
                    celda_num.value = i
                    celda_num.alignment = center_align
                
                # Nombre (Columna 2 / B)
                celda_nombre = ws.cell(row=fila_actual, column=2)
                if type(celda_nombre).__name__ != 'MergedCell':
                    celda_nombre.value = nombre
                    celda_nombre.alignment = left_align

                # Sexo (Columna 3 / C)
                celda_sexo = ws.cell(row=fila_actual, column=3)
                if type(celda_sexo).__name__ != 'MergedCell':
                    celda_sexo.value = letra_sexo
                    celda_sexo.alignment = center_align

                # Asistencias e inasistencias por semana y día (A partir de Columna 4 / D)
                total_presentes = 0
                for sem, start_col in semanas_cols.items():
                    for offset, dia in enumerate(dias):
                        presente = mapa_asis.get(ced, {}).get(sem, {}).get(dia, False)
                        celda = ws.cell(row=fila_actual, column=start_col + offset)
                        
                        if type(celda).__name__ != 'MergedCell':
                            if presente:
                                celda.value = "P"
                                total_presentes += 1
                            else:
                                celda.value = "X"
                            celda.alignment = center_align

                # Totalizador automático de días asistidos (Columna 24 / X)
                celda_total = ws.cell(row=fila_actual, column=24)
                if type(celda_total).__name__ != 'MergedCell':
                    celda_total.value = total_presentes
                    celda_total.alignment = center_align

                # Porcentaje de asistencia del alumno en el mes (Columna 25 / Y)
                celda_porc = ws.cell(row=fila_actual, column=25)
                if type(celda_porc).__name__ != 'MergedCell':
                    porcentaje_calculado = (total_presentes / total_dias_mes) if total_dias_mes > 0 else 0
                    celda_porc.value = porcentaje_calculado
                    celda_porc.number_format = '0.0%'
                    celda_porc.alignment = center_align
                    celda_porc.border = borde_fino
                
                fila_actual += 1

            # 7. Colocar totales de Varones y Hembras evadiendo celdas combinadas ("MergedCell")
            celda_varones_set = False
            celda_hembras_set = False
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=15):
                for cell in row:
                    if type(cell).__name__ == 'MergedCell':
                        continue
                        
                    if cell.value and isinstance(cell.value, str):
                        texto = cell.value.strip().lower()
                        
                        if "varon" in texto or "varón" in texto:
                            col_dest = cell.column + 1
                            while type(ws.cell(row=cell.row, column=col_dest)).__name__ == 'MergedCell':
                                col_dest += 1
                                
                            celda_dest = ws.cell(row=cell.row, column=col_dest)
                            celda_dest.value = total_varones
                            celda_dest.alignment = center_align
                            celda_varones_set = True
                            
                        elif "hembra" in texto:
                            col_dest = cell.column + 1
                            while type(ws.cell(row=cell.row, column=col_dest)).__name__ == 'MergedCell':
                                col_dest += 1
                                
                            celda_dest = ws.cell(row=cell.row, column=col_dest)
                            celda_dest.value = total_hembras
                            celda_dest.alignment = center_align
                            celda_hembras_set = True

            if not celda_varones_set:
                ws.cell(row=fila_actual + 1, column=2, value="Total Varones:").alignment = left_align
                ws.cell(row=fila_actual + 1, column=3, value=total_varones).alignment = center_align
            if not celda_hembras_set:
                ws.cell(row=fila_actual + 2, column=2, value="Total Hembras:").alignment = left_align
                ws.cell(row=fila_actual + 2, column=3, value=total_hembras).alignment = center_align

            try:
                window = webview.windows[0]
                
                # Limpiamos los textos para que el explorador de archivos no tenga problemas (reemplazar espacios con guiones bajos)
                grado_limpio = grado.replace(" ", "_")
                seccion_limpia = seccion.replace(" ", "_")
                
                # Saneamos el nombre de cualquier carácter no válido para rutas como '/', o '\'
                nombre_sugerido = f"Asistencias_Detalladas_{grado_limpio}_Sec_{seccion_limpia}_{mes}.xlsx"
                nombre_sugerido = nombre_sugerido.replace("/", "-").replace("\\", "-")
                
                if not nombre_sugerido.endswith('.xlsx'):
                    nombre_sugerido += '.xlsx'
                
                result = window.create_file_dialog(
                    webview.SAVE_DIALOG, 
                    directory='', 
                    save_filename=nombre_sugerido
                )

                if result:
                    ruta_guardado = result[0] if isinstance(result, tuple) else result
                    
                    if not ruta_guardado.endswith('.xlsx'):
                        ruta_guardado += '.xlsx'
                    
                    wb.save(ruta_guardado)
                    return {"status": "success", "ruta": ruta_guardado}
                else:
                    return {"status": "error", "message": "Guardado cancelado por el usuario."}
                    
            except Exception as w_err:
                return {"status": "error", "message": f"Error abriendo ventana de guardado: {str(w_err)}"}

        except Exception as e:
            return {"status": "error", "message": str(e)}