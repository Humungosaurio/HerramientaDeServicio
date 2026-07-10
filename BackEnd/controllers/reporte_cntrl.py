import os
import webview
import openpyxl

class ReporteController:
    def __init__(self):
        pass

    def generar_excel_desde_plantilla(self, datos_estudiantes, nombre_archivo):
        try:
            # Subimos un nivel porque ahora estamos dentro de la carpeta 'controllers'
            base_dir = os.path.dirname(os.path.dirname(__file__))
            ruta_plantilla = os.path.join(base_dir, "plantilla_matricula.xlsx")
            
            if not os.path.exists(ruta_plantilla):
                return {"status": "error", "message": "No se encontró el archivo plantilla_matricula.xlsx en el backend."}

            window = webview.windows[0]
            result = window.create_file_dialog(
                webview.SAVE_DIALOG,
                directory='',
                save_filename=f"{nombre_archivo}.xlsx",
                file_types=('Archivos Excel (*.xlsx)', 'Todos los archivos (*.*)')
            )

            if not result:
                return {"status": "cancelado"}

            ruta_guardado = result[0]

            # Cargar plantilla preservando imágenes (Pillow requerido)
            wb = openpyxl.load_workbook(ruta_plantilla)
            ws = wb.active

            # =========================================================
            # FUNCIÓN AUXILIAR PARA CELDAS COMBINADAS
            # =========================================================
            def escribir_celda_segura(fila, col, valor):
                try:
                    ws.cell(row=fila, column=col, value=valor)
                except AttributeError:
                    celda = ws.cell(row=fila, column=col)
                    for rango in ws.merged_cells.ranges:
                        if celda.coordinate in rango:
                            ws.cell(row=rango.min_row, column=rango.min_col, value=valor)
                            break

            # =========================================================
            # COORDENADAS EXACTAS DEL FORMATO DEA RR-DEA-07-04
            # =========================================================
            fila_superior = 18 
            fila_inferior = 40 

            for i, est in enumerate(datos_estudiantes):
                # --- BLOQUE SUPERIOR ---
                escribir_celda_segura(fila_superior, 1, i + 1)
                
                cedula = est.get('cedulaEscolar') or est.get('cedula_estudiantil') or ''
                escribir_celda_segura(fila_superior, 2, cedula)
                
                lugar_nac = est.get('lugarNacimiento') or est.get('lugar_nacimiento') or ''
                escribir_celda_segura(fila_superior, 5, str(lugar_nac).upper())

                ef = est.get('entidadFederal') or est.get('ef') or est.get('est_ef') or ''
                escribir_celda_segura(fila_superior, 10, str(ef).upper())

                genero = est.get('genero') or est.get('est_genero') or ''
                escribir_celda_segura(fila_superior, 11, genero[0].upper() if genero else '')

                fecha_nac = est.get('fechaNacimiento') or est.get('est_fecha_nacimiento') or ''
                if fecha_nac:
                    sep = '-' if '-' in fecha_nac else '/'
                    partes = fecha_nac.split(sep)
                    if len(partes) == 3:
                        if len(partes[0]) == 4: 
                            dia, mes, anio = partes[2], partes[1], partes[0]
                        else:
                            dia, mes, anio = partes[0], partes[1], partes[2]
                            
                        escribir_celda_segura(fila_superior, 12, dia)  
                        escribir_celda_segura(fila_superior, 13, mes)  
                        escribir_celda_segura(fila_superior, 14, anio) 

                # --- BLOQUE INFERIOR ---
                escribir_celda_segura(fila_inferior, 1, i + 1)
                
                apellidos = est.get('apellidos', '')
                nombres = est.get('nombres', '')
                
                if not apellidos and not nombres:
                    nombre_completo = str(est.get('nombre') or est.get('est_nombre') or '').strip().upper()
                    partes_nombre = nombre_completo.split()
                    
                    if len(partes_nombre) >= 3:
                        apellidos = " ".join(partes_nombre[:2])
                        nombres = " ".join(partes_nombre[2:])
                    elif len(partes_nombre) == 2:
                        apellidos = partes_nombre[0]
                        nombres = partes_nombre[1]
                    else:
                        apellidos = nombre_completo
                        nombres = ""

                escribir_celda_segura(fila_inferior, 2, str(apellidos).upper())
                escribir_celda_segura(fila_inferior, 11, str(nombres).upper())

                fila_superior += 1
                fila_inferior += 1

            wb.save(ruta_guardado)
            return {"status": "success"}

        except Exception as e:
            print(f"❌ Error al generar Excel: {e}")
            return {"status": "error", "message": str(e)}