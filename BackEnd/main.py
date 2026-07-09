import os
import sqlite3
import webview
import openpyxl
from controllers.register_cntrl import AcademicController
from controllers.asis_det_cntrl import Asis_Det_Controller
from controllers.personal_cntrl import PersonalController
from controllers.inventario_cntrl import InventarioController

def inicializar_base_de_datos(db_path):
    db_dir = os.path.dirname(db_path)
    if not os.path.exists(db_dir):
        os.makedirs(db_dir, exist_ok=True)
        print(f"📁 Directorio creado: {db_dir}")

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    conn.execute("PRAGMA foreign_keys = ON;")

    esquema_sql = """
    CREATE TABLE IF NOT EXISTS salones (
        salon_id INTEGER PRIMARY KEY AUTOINCREMENT,
        grado TEXT NOT NULL,
        turno TEXT NOT NULL,
        seccion TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS representante (
        representante_ci INTEGER PRIMARY KEY,
        nombre TEXT NOT NULL,
        direccion TEXT,
        fecha_nacimiento TEXT,
        grado_educacion TEXT,
        trabaja INTEGER,
        direccion_trabajo TEXT,
        parentesco TEXT,
        lugar_nacimiento TEXT,
        telefono TEXT,
        correo TEXT
    );

    CREATE TABLE IF NOT EXISTS rep_responsable (
        re_inst_ci INTEGER PRIMARY KEY,
        nombre TEXT NOT NULL,
        direccion TEXT,
        fecha_nacimiento TEXT,
        grado_educacion TEXT,
        trabaja INTEGER,
        direccion_trabajo TEXT,
        parentesco TEXT,
        lugar_nacimiento TEXT,
        telefono TEXT,
        correo TEXT
    );

    CREATE TABLE IF NOT EXISTS Estudiante (
        cedula_estudiantil INTEGER PRIMARY KEY,
        est_nombre TEXT NOT NULL,
        est_direccion TEXT,
        est_genero TEXT NOT NULL,
        neurodiversidad TEXT,
        est_fecha_nacimiento TEXT,
        re_inst_ci INTEGER,
        representante_ci INTEGER,
        salon_id INTEGER NOT NULL,
        tipo_sangre TEXT,
        talla_mono TEXT,
        talla_camisa TEXT,
        talla_calzado TEXT,
        FOREIGN KEY (salon_id) REFERENCES salones(salon_id) ON DELETE RESTRICT,
        FOREIGN KEY (representante_ci) REFERENCES representante(representante_ci) ON DELETE SET NULL,
        FOREIGN KEY (re_inst_ci) REFERENCES rep_responsable(re_inst_ci) ON DELETE SET NULL
    );

    CREATE TABLE IF NOT EXISTS asistencias (
        asistencia_id INTEGER PRIMARY KEY AUTOINCREMENT,
        cedula_estudiantil INTEGER NOT NULL,
        mes TEXT NOT NULL,
        semana TEXT NOT NULL,
        dia_semana TEXT NOT NULL,
        estado TEXT NOT NULL,
        UNIQUE(cedula_estudiantil, mes, semana, dia_semana),
        FOREIGN KEY (cedula_estudiantil) REFERENCES Estudiante(cedula_estudiantil) ON DELETE CASCADE
    );

        CREATE TABLE IF NOT EXISTS personal (
        cedula_trabajador INTEGER PRIMARY KEY,
        nombre TEXT NOT NULL,
        cargo TEXT NOT NULL,
        turno TEXT,
        horas_administrativas INTEGER
    );

    CREATE TABLE IF NOT EXISTS asistencias_personal (
        asistencia_id INTEGER PRIMARY KEY AUTOINCREMENT,
        cedula_trabajador INTEGER NOT NULL,
        mes TEXT NOT NULL,
        semana TEXT NOT NULL,
        dia_semana TEXT NOT NULL,
        estado TEXT NOT NULL,
        FOREIGN KEY (cedula_trabajador) REFERENCES personal(cedula_trabajador) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS inventario_mobiliario (
        mobiliario_id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        codigo_bien TEXT UNIQUE,
        cantidad INTEGER DEFAULT 0,
        activo TEXT,
        comentarios TEXT
    );
    """

    try:
        cursor.executescript(esquema_sql)
        cursor.execute("SELECT COUNT(*) FROM salones")
        if cursor.fetchone()[0] == 0:
            print("🏫 Tabla 'salones' vacía. Precargando la configuración institucional...")
            salones_iniciales = [
                ('Maternal', 'Mañana', 'A'), ('Maternal', 'Tarde', 'B'),
                ('1er Nivel', 'Mañana', 'A'), ('1er Nivel', 'Mañana', 'B'),
                ('1er Nivel', 'Tarde', 'C'), ('1er Nivel', 'Tarde', 'D'),
                ('2do Nivel', 'Mañana', 'A'), ('2do Nivel', 'Mañana', 'B'),
                ('2do Nivel', 'Tarde', 'C'), ('2do Nivel', 'Tarde', 'D'),
                ('3er Nivel', 'Mañana', 'A'), ('3er Nivel', 'Tarde', 'B'), ('3er Nivel', 'Tarde', 'C')
            ]
            cursor.executemany(
                "INSERT INTO salones (grado, turno, seccion) VALUES (?, ?, ?)", 
                salones_iniciales
            )
            print(f"✅ Se han registrado {len(salones_iniciales)} aulas correctamente.")

        conn.commit()
        print("✅ Estructura de base de datos verificada y sincronizada con los controladores.")
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Error al inicializar las tablas en main.py: {e}")
    finally:
        conn.close()


class SistemaAPI:
    def __init__(self):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.db_path = os.path.join(base_dir, "database", "data", "Control_Estudiantil.db")
        
        inicializar_base_de_datos(self.db_path)
        
        self.controlador_academico = AcademicController(self.db_path)
        self.controlador_asistencia = Asis_Det_Controller(self.db_path)
        self.controlador_personal = PersonalController(self.db_path)
        self.controlador_inventario = InventarioController(self.db_path)

    # =========================================================
    # PUENTES ACADÉMICOS Y ESTUDIANTILES
    # =========================================================
    def registrar_estudiante_completo(self, data):
        return self.controlador_academico.registrar_estudiante_completo(data)

    # NUEVO PUENTE: Permite al frontend pedir la lista al recargar
    def obtener_estudiantes(self):
        return self.controlador_academico.obtener_estudiantes()

    def cargar_matriz_asistencia(self, parametros):
        return self.controlador_asistencia.cargar_matriz_asistencia(parametros)

    def guardar_asistencias(self, data):
        return self.controlador_asistencia.guardar_asistencias(data)

    def obtener_resumen_global(self, parametros):
        return self.controlador_asistencia.obtener_resumen_global(parametros)
    
    # =========================================================
    # PUENTES PARA EL PERSONAL
    # =========================================================
    def registrar_trabajador(self, data):
        return self.controlador_personal.registrar_trabajador(data)

    def cargar_matriz_asistencia_personal(self, parametros):
        return self.controlador_personal.cargar_matriz_asistencia_personal(parametros)

    def guardar_asistencias_personal(self, data):
        return self.controlador_personal.guardar_asistencias_personal(data)
    
    # =========================================================
    # PUENTES PARA EL INVENTARIO
    # =========================================================
    def cargar_inventario(self):
        return self.controlador_inventario.cargar_inventario()

    def guardar_inventario_masivo(self, lista_bienes):
        return self.controlador_inventario.guardar_inventario_masivo(lista_bienes)

    def eliminar_articulo_inventario(self, id_articulo):
        return self.controlador_inventario.eliminar_articulo(id_articulo)
    def generar_excel_desde_plantilla(self, datos_estudiantes, nombre_archivo):
        try:
            ruta_plantilla = os.path.join(os.path.dirname(__file__), "plantilla_matricula.xlsx")
            
            if not os.path.exists(ruta_plantilla):
                return {"status": "error", "message": "No se encontró el archivo plantilla_matricula.xlsx en el backend."}

            window = webview.windows[0]
            result = window.create_file_dialog(
                webview.SAVE_DIALOG,
                directory='',
                save_filename=f"{nombre_archivo}.xlsx",
                file_types=('Archivos Excel (*.xlsx)', 'Todos los archivos (*.*)')
            )

            if not result:
                return {"status": "cancelado"}

            ruta_guardado = result[0]

            # Cargar plantilla preservando imágenes (Pillow requerido)
            wb = openpyxl.load_workbook(ruta_plantilla)
            ws = wb.active

            # =========================================================
            # FUNCIÓN AUXILIAR PARA CELDAS COMBINADAS
            # =========================================================
            def escribir_celda_segura(fila, col, valor):
                try:
                    ws.cell(row=fila, column=col, value=valor)
                except AttributeError:
                    celda = ws.cell(row=fila, column=col)
                    for rango in ws.merged_cells.ranges:
                        if celda.coordinate in rango:
                            ws.cell(row=rango.min_row, column=rango.min_col, value=valor)
                            break

            # =========================================================
            # COORDENADAS EXACTAS DEL FORMATO DEA RR-DEA-07-04
            # =========================================================
            fila_superior = 18 # Fila donde inicia el estudiante Nº 01 en la tabla superior
            fila_inferior = 40 # Fila donde inicia el estudiante Nº 01 en la tabla inferior

            for i, est in enumerate(datos_estudiantes):
                # -------------------------------------------------------------
                # 1. BLOQUE SUPERIOR (Cédula, Lugar, Sexo, Fecha - Fila 18+)
                # -------------------------------------------------------------
                # Columna A (1): Nº de lista
                escribir_celda_segura(fila_superior, 1, i + 1)
                
                # Columna B (2): Cédula de Identidad o Cédula Escolar
                cedula = est.get('cedulaEscolar') or est.get('cedula_estudiantil') or ''
                escribir_celda_segura(fila_superior, 2, cedula)
                
                # Columna E (5): Lugar de Nacimiento
                lugar_nac = est.get('lugarNacimiento') or est.get('lugar_nacimiento') or ''
                escribir_celda_segura(fila_superior, 5, str(lugar_nac).upper())

                # Columna J (10): Entidad Federal (EF) - Ej: CARABOBO
                ef = est.get('entidadFederal') or est.get('ef') or est.get('est_ef') or ''
                escribir_celda_segura(fila_superior, 10, str(ef).upper())

                # Columna K (11): Sexo (M o F)
                genero = est.get('genero') or est.get('est_genero') or ''
                escribir_celda_segura(fila_superior, 11, genero[0].upper() if genero else '')

                # Columna L (12), M (13), N (14): Fecha de Nacimiento (Día, Mes, Año)
                fecha_nac = est.get('fechaNacimiento') or est.get('est_fecha_nacimiento') or ''
                if fecha_nac:
                    sep = '-' if '-' in fecha_nac else '/'
                    partes = fecha_nac.split(sep)
                    if len(partes) == 3:
                        # Detecta si viene YYYY-MM-DD y lo ordena a Día, Mes, Año
                        if len(partes[0]) == 4: 
                            dia, mes, anio = partes[2], partes[1], partes[0]
                        else:
                            dia, mes, anio = partes[0], partes[1], partes[2]
                            
                        escribir_celda_segura(fila_superior, 12, dia)  # Columna L: Día
                        escribir_celda_segura(fila_superior, 13, mes)  # Columna M: Mes
                        escribir_celda_segura(fila_superior, 14, anio) # Columna N: Año

                # -------------------------------------------------------------
                # 2. BLOQUE INFERIOR (Apellidos y Nombres - Fila 40+)
                # -------------------------------------------------------------
                # Columna A (1): Nº de lista en tabla inferior
                escribir_celda_segura(fila_inferior, 1, i + 1)
                
                # Lógica para separar Apellidos y Nombres si vienen juntos en 'nombre'
                apellidos = est.get('apellidos', '')
                nombres = est.get('nombres', '')
                
                if not apellidos and not nombres:
                    nombre_completo = str(est.get('nombre') or est.get('est_nombre') or '').strip().upper()
                    partes_nombre = nombre_completo.split()
                    
                    if len(partes_nombre) >= 3:
                        # Asume 2 apellidos y el resto son nombres (Ej: PÉREZ GONZÁLEZ JUAN CARLOS)
                        apellidos = " ".join(partes_nombre[:2])
                        nombres = " ".join(partes_nombre[2:])
                    elif len(partes_nombre) == 2:
                        apellidos = partes_nombre[0]
                        nombres = partes_nombre[1]
                    else:
                        apellidos = nombre_completo
                        nombres = ""

                # Columna B (2): Apellidos (ocupa de B hasta J)
                escribir_celda_segura(fila_inferior, 2, str(apellidos).upper())
                
                # Columna K (11): Nombres (ocupa de K en adelante)
                escribir_celda_segura(fila_inferior, 11, str(nombres).upper())

                # Avanzamos a la siguiente fila en ambos bloques
                fila_superior += 1
                fila_inferior += 1

            wb.save(ruta_guardado)
            return {"status": "success"}

        except Exception as e:
            print(f"❌ Error al generar Excel: {e}")
            return {"status": "error", "message": str(e)}
def iniciar_aplicacion():
    api_global = SistemaAPI()
    window = webview.create_window(
        title="Registro Administrativo - Simoncito",
        url="http://localhost:5173",
        js_api=api_global,
        width=1200, height=700, resizable=True, min_size=(1024, 600)
    )
    webview.start(debug=True)

if __name__ == "__main__":
    iniciar_aplicacion()